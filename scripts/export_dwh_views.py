"""Export DWH views to an Excel workbook for downstream spreadsheet use."""

from __future__ import annotations

import argparse
import sys
from datetime import UTC, datetime
from pathlib import Path
import shutil

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pandas import DatetimeTZDtype
from sqlalchemy import text

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from db.session import engine

EXPORT_DIR = PROJECT_ROOT / "exports"

VIEW_SHEETS: list[tuple[str, str]] = [
    ("dwh.v_top_news_week_export", "top_week"),
    ("dwh.v_top_news_month_export", "top_month"),
    ("dwh.v_news_week_export", "news_week"),
    ("dwh.v_news_month_export", "news_month"),
    ("dwh.v_news_6_months_export", "news_6_months"),
    ("dwh.v_news_all_export", "news_all"),
]

DEA_VIEW_SHEETS: list[tuple[str, str, str]] = [
    ("dea.v_kpi_overview", "dea_kpis", "ORDER BY metric"),
    (
        "dea.v_company_intelligence",
        "dea_company",
        "ORDER BY high_priority_last_30_days DESC, articles_last_30_days DESC, company_name",
    ),
    (
        "dea.v_topic_signal_heatmap",
        "dea_topic_signal",
        "ORDER BY high_priority_last_30_days DESC, articles_last_30_days DESC, key_topic, signal_type",
    ),
    (
        "dea.v_executive_news_feed",
        "dea_exec_feed",
        "ORDER BY priority_score DESC NULLS LAST, published_date DESC NULLS LAST, company_name, title",
    ),
]

COMPANY_COLORS = {
    "ALLIANCE HEALTHCARE": "D8EAF7",
    "ASTERA": "DFF2E1",
    "BIOCODEX": "FADBD8",
    "CEVA SANTE": "E8DAEF",
    "DELPHARM": "FCF3CF",
    "EUROFINS": "D6EAF8",
    "FAREVA": "F5C6C6",
    "GALDERMA": "D5F5E3",
    "HAELON": "FDEBD0",
    "IPSEN": "D4E6F1",
    "LILLY": "FADBD8",
    "MODERNA": "E8DAEF",
    "OPELLA": "F9E79F",
    "OXIPHARM": "D1F2EB",
    "PFIZER": "CFE2FF",
    "PIERRE FABRE": "FDEDEC",
    "SANOFI": "D6EAF8",
    "SEBIA": "EBDEF0",
    "SERVIER": "FAD7A0",
    "STAGO": "D5F5E3",
    "VIATRIS": "EAECEE",
    "VIRBAC": "D4EFDF",
}


def fetch_view(
    view_name: str,
    order_clause: str = "ORDER BY priority_score DESC NULLS LAST, published_date DESC NULLS LAST, company_name, title",
) -> pd.DataFrame:
    with engine.begin() as connection:
        frame = pd.read_sql_query(
            text(f"SELECT * FROM {view_name} {order_clause}"),
            connection,
        )
    return frame


def normalize_datetimes(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = frame.copy()
    for column in normalized.columns:
        if isinstance(normalized[column].dtype, DatetimeTZDtype):
            normalized[column] = normalized[column].dt.tz_convert(None)
        elif pd.api.types.is_datetime64_any_dtype(normalized[column]):
            normalized[column] = normalized[column].dt.tz_localize(None)
    return normalized


def build_overview_frames(frames: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    all_frame = frames["news_all"]
    week_frame = frames["news_week"]
    month_frame = frames["news_month"]

    overview = pd.DataFrame(
        [
            {"metric": "generated_at_utc", "value": datetime.now(UTC).isoformat()},
            {"metric": "all_articles", "value": len(all_frame.index)},
            {"metric": "week_articles", "value": len(week_frame.index)},
            {"metric": "month_articles", "value": len(month_frame.index)},
            {
                "metric": "high_priority_articles",
                "value": int(all_frame["priority_score"].ge(75).sum()) if "priority_score" in all_frame else 0,
            },
        ]
    )

    company_counts = (
        month_frame.groupby("company_name", dropna=False)
        .size()
        .reset_index(name="article_count")
        .sort_values(["article_count", "company_name"], ascending=[False, True])
    )

    topic_counts = (
        all_frame.groupby("key_topic", dropna=False)
        .size()
        .reset_index(name="article_count")
        .sort_values(["article_count", "key_topic"], ascending=[False, True])
    )

    signal_counts = (
        all_frame.groupby("signal_type", dropna=False)
        .size()
        .reset_index(name="article_count")
        .sort_values(["article_count", "signal_type"], ascending=[False, True])
    )

    return {
        "overview": overview,
        "company_summary": company_counts,
        "topic_summary": topic_counts,
        "signal_summary": signal_counts,
    }


def apply_company_row_colors(worksheet, frame: pd.DataFrame) -> None:
    if "company_name" not in frame.columns:
        return

    company_col_idx = list(frame.columns).index("company_name") + 1
    for row_idx in range(2, worksheet.max_row + 1):
        company_name = worksheet.cell(row=row_idx, column=company_col_idx).value
        if not company_name:
            continue
        fill_color = COMPANY_COLORS.get(str(company_name).upper())
        if not fill_color:
            continue
        fill = PatternFill(fill_type="solid", fgColor=fill_color)
        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = fill


def style_sheet(writer: pd.ExcelWriter, sheet_name: str, frame: pd.DataFrame, color_by_company: bool = False) -> None:
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    wrap_columns = {"title", "article_summary", "business_impact", "url"}
    preferred_widths = {
        "company_name": 22,
        "published_date": 20,
        "priority_score": 14,
        "title": 42,
        "article_summary": 88,
        "key_topic": 18,
        "business_impact": 52,
        "geography": 18,
        "signal_type": 20,
        "url": 44,
        "summary_status": 14,
        "metric": 28,
        "metric_label": 52,
        "value": 22,
        "numeric_value": 16,
        "metric_value": 16,
        "article_count": 14,
        "articles_all": 12,
        "articles_last_7_days": 18,
        "articles_last_30_days": 18,
        "articles_last_6_months": 20,
        "high_priority_last_30_days": 24,
        "avg_priority_last_30_days": 22,
        "top_priority_score_30_days": 24,
        "leading_topic_last_30_days": 26,
        "leading_signal_last_30_days": 26,
        "attention_level": 16,
        "priority_band": 14,
    }

    for idx, column_name in enumerate(frame.columns, start=1):
        column_letter = get_column_letter(idx)
        width = preferred_widths.get(column_name, max(len(str(column_name)) + 2, 14))
        worksheet.column_dimensions[column_letter].width = width

        for cell in worksheet[column_letter]:
            cell.alignment = Alignment(vertical="top", wrap_text=column_name in wrap_columns)

    if color_by_company:
        apply_company_row_colors(worksheet, frame)


def export_workbook(output_dir: Path) -> tuple[Path, Path, dict[str, int]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    latest_path = output_dir / "lifescience_watch_news_latest.xlsx"
    archive_path = output_dir / f"lifescience_watch_news_{timestamp}.xlsx"
    sheet_counts: dict[str, int] = {}

    frames: dict[str, pd.DataFrame] = {}
    for view_name, sheet_name in VIEW_SHEETS:
        frames[sheet_name] = normalize_datetimes(fetch_view(view_name))

    dea_frames: dict[str, pd.DataFrame] = {}
    for view_name, sheet_name, order_clause in DEA_VIEW_SHEETS:
        dea_frames[sheet_name] = normalize_datetimes(fetch_view(view_name, order_clause))

    summary_frames = build_overview_frames(frames)

    with pd.ExcelWriter(latest_path, engine="openpyxl") as writer:
        for sheet_name, frame in summary_frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            style_sheet(writer, sheet_name, frame, color_by_company=sheet_name == "company_summary")
            sheet_counts[sheet_name] = len(frame.index)

        for sheet_name, frame in dea_frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            style_sheet(writer, sheet_name, frame, color_by_company="company_name" in frame.columns)
            sheet_counts[sheet_name] = len(frame.index)

        for _, sheet_name in VIEW_SHEETS:
            frame = frames[sheet_name]
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            style_sheet(writer, sheet_name, frame, color_by_company=True)
            sheet_counts[sheet_name] = len(frame.index)

        metadata = pd.DataFrame(
            [
                {"key": "generated_at_utc", "value": datetime.now(UTC).isoformat()},
                {"key": "source_database", "value": engine.url.render_as_string(hide_password=True)},
                {"key": "views_exported", "value": ", ".join(view for view, _ in VIEW_SHEETS)},
                {"key": "dea_views_exported", "value": ", ".join(view for view, _, _ in DEA_VIEW_SHEETS)},
            ]
        )
        metadata.to_excel(writer, sheet_name="export_info", index=False)
        style_sheet(writer, "export_info", metadata)

    shutil.copy2(latest_path, archive_path)
    return latest_path, archive_path, sheet_counts


def main() -> int:
    parser = argparse.ArgumentParser(description="Export DWH views to Excel.")
    parser.add_argument(
        "--output-dir",
        default=str(EXPORT_DIR),
        help="Destination folder for the workbook export.",
    )
    args = parser.parse_args()

    latest_path, archive_path, sheet_counts = export_workbook(Path(args.output_dir))
    count_summary = ", ".join(f"{sheet}={count}" for sheet, count in sheet_counts.items())
    print(f"latest={latest_path}")
    print(f"archive={archive_path}")
    print(f"counts={count_summary}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
